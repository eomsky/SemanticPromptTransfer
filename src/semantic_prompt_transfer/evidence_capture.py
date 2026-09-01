from __future__ import annotations

import io
import re
import textwrap
from pathlib import Path
from typing import Any

from .colab_runtime import EphemeralColabRuntime


class EvidenceCaptureService:
    """Render a bounded, case-scoped source excerpt for one cited evidence ID."""

    def __init__(self, runtime: EphemeralColabRuntime) -> None:
        self.runtime = runtime

    def _document(self, tenant_id: str, case_id: str, evidence: dict[str, Any]):
        document_id = str(evidence.get("document_id") or "")
        document = self.runtime.registry.get_document(tenant_id, case_id, document_id)
        if not document.storage_uri:
            raise RuntimeError("evidence source file is unavailable")
        source = Path(document.storage_uri).expanduser().resolve()
        if not source.is_relative_to(self.runtime.root.resolve()) or not source.is_file():
            raise PermissionError("evidence source is outside the case runtime")
        return document, source

    @staticmethod
    def _font(size: int, *, bold: bool = False):
        from PIL import ImageFont

        candidates = (
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc" if bold else "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
            "/usr/local/share/fonts/spt_v022/NanumGothicBold.ttf" if bold else "/usr/local/share/fonts/spt_v022/NanumGothic.ttf",
            "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf" if bold else "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        )
        for candidate in candidates:
            if Path(candidate).is_file():
                return ImageFont.truetype(candidate, size=size)
        return ImageFont.load_default()

    @staticmethod
    def _png(image) -> bytes:
        output = io.BytesIO()
        image.save(output, format="PNG", optimize=True)
        return output.getvalue()

    def describe(
        self,
        tenant_id: str,
        case_id: str,
        evidence_id: str,
        evidence: dict[str, Any],
    ) -> dict[str, Any]:
        document, source = self._document(tenant_id, case_id, evidence)
        metadata = dict(evidence.get("metadata") or {})
        suffix = source.suffix.lower()
        if suffix == ".pdf":
            kind = "pdf"
            location = f"{int(evidence.get('page') or 1)}페이지"
        elif suffix == ".xlsx":
            kind = "xlsx"
            sheet_name, cell_range = self._xlsx_location(metadata)
            location = f"{sheet_name or '시트'} · {cell_range or '범위'}"
        else:
            kind = "text"
            location = str(metadata.get("source_location") or "원문")
        return {
            "evidence_id": evidence_id,
            "kind": kind,
            "source_tier": int(evidence.get("source_tier") or 3),
            "filename": document.filename,
            "location": location,
            "excerpt": str(evidence.get("content") or ""),
            "capture_available": True,
        }

    def capture_png(
        self,
        tenant_id: str,
        case_id: str,
        evidence: dict[str, Any],
    ) -> bytes:
        _, source = self._document(tenant_id, case_id, evidence)
        if source.suffix.lower() == ".pdf":
            return self._pdf(source, evidence)
        if source.suffix.lower() == ".xlsx":
            return self._xlsx(source, evidence)
        return self._text(evidence)

    def _pdf(self, source: Path, evidence: dict[str, Any]) -> bytes:
        import fitz
        from PIL import Image, ImageDraw

        metadata = dict(evidence.get("metadata") or {})
        page_number = max(1, int(evidence.get("page") or 1))
        document = fitz.open(str(source))
        try:
            if page_number > document.page_count:
                raise ValueError("evidence page is outside the PDF")
            page = document[page_number - 1]
            raw_bbox = metadata.get("bbox")
            bbox = fitz.Rect(*(float(v) for v in raw_bbox)) if isinstance(raw_bbox, (list, tuple)) and len(raw_bbox) == 4 else None
            boxes = []
            for value in metadata.get("highlight_bboxes") or []:
                if isinstance(value, (list, tuple)) and len(value) == 4:
                    boxes.append(fitz.Rect(*(float(x) for x in value)))
            if bbox is not None and not boxes: boxes = [bbox]

            if bbox is None:
                clip = page.rect
            elif metadata.get("logical_table_id"):
                clip = fitz.Rect(
                    page.rect.x0 + 16.0, max(page.rect.y0, bbox.y0 - max(110.0, page.rect.height * 0.24)),
                    page.rect.x1 - 16.0, min(page.rect.y1, bbox.y1 + max(75.0, page.rect.height * 0.15)),
                )
            else:
                desired_width = max(bbox.width + 180.0, page.rect.width * 0.78)
                center = (bbox.x0 + bbox.x1) / 2
                x0 = max(page.rect.x0, center - desired_width / 2); x1 = min(page.rect.x1, center + desired_width / 2)
                if x1 - x0 < desired_width: x0 = max(page.rect.x0, x1 - desired_width)
                pad_y = max(90.0, bbox.height * 2.5)
                clip = fitz.Rect(x0, max(page.rect.y0, bbox.y0 - pad_y), x1, min(page.rect.y1, bbox.y1 + pad_y))

            scale = 1.45
            pixmap = page.get_pixmap(matrix=fitz.Matrix(scale, scale), clip=clip, alpha=False)
            image = Image.open(io.BytesIO(pixmap.tobytes("png"))).convert("RGB")
            overlay = ImageDraw.Draw(image, "RGBA")
            for box in boxes:
                left = max(0, int((box.x0 - clip.x0) * scale)); top = max(0, int((box.y0 - clip.y0) * scale))
                right = min(image.width - 1, int((box.x1 - clip.x0) * scale)); bottom = min(image.height - 1, int((box.y1 - clip.y0) * scale))
                if right > left and bottom > top:
                    overlay.rectangle((left, top, right, bottom), fill=(255, 213, 0, 48), outline=(255, 150, 0, 255), width=4)
            return self._png(image)
        finally:
            document.close()

    def _xlsx(self, source: Path, evidence: dict[str, Any]) -> bytes:
        from openpyxl import load_workbook
        from openpyxl.cell.cell import MergedCell
        from openpyxl.styles.numbers import is_date_format
        from openpyxl.utils import get_column_letter
        from openpyxl.utils.cell import range_boundaries
        from PIL import Image, ImageDraw

        metadata = dict(evidence.get("metadata") or {})
        sheet_name, cell_range = self._xlsx_location(metadata)
        workbook = load_workbook(source, data_only=True, read_only=False)
        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError("evidence sheet is unavailable")
            sheet = workbook[sheet_name]
            if cell_range.startswith("ROW:"):
                row_number = max(1, int(cell_range.split(":", 1)[1]))
                cell_range = f"A{row_number}:{get_column_letter(min(max(1, sheet.max_column), 12))}{row_number}"
            highlight_ranges = [cell_range]
            for extra in metadata.get("highlight_cell_ranges") or []:
                value = str(extra or "").strip()
                if value and value not in highlight_ranges and not value.startswith("ROW:"):
                    highlight_ranges.append(value)
            bounds = [range_boundaries(value) for value in highlight_ranges]
            min_col = min(value[0] for value in bounds); min_row = min(value[1] for value in bounds)
            max_col = max(value[2] for value in bounds); max_row = max(value[3] for value in bounds)
            first_row = max(1, min_row - 10)
            for candidate in range(min_row - 1, first_row - 1, -1):
                values = [sheet.cell(candidate, c).value for c in range(max(1, min_col - 2), min(sheet.max_column, max_col + 2) + 1)]
                nonempty = [value for value in values if value not in (None, "")]
                if len(nonempty) >= 2 and any(isinstance(value, str) for value in nonempty):
                    first_row = candidate
            last_row = min(sheet.max_row, max_row + 5)
            first_col = 1 if sheet.max_column <= 14 else max(1, min_col - 3)
            last_col = min(sheet.max_column, max(max_col + 3, first_col + 8))
            scale = 2
            row_header = 48 * scale
            column_header = 28 * scale

            def column_pixels(column: int) -> int:
                dimension = sheet.column_dimensions[get_column_letter(column)]
                width = float(dimension.width or 8.43)
                return int(max(58, min(360, width * 9.2 + 14)) * scale)

            def row_pixels(row: int) -> int:
                height = float(sheet.row_dimensions[row].height or 22)
                return int(max(30, min(140, height * 96 / 72)) * scale)

            column_widths = [column_pixels(column) for column in range(first_col, last_col + 1)]
            row_heights = [row_pixels(row) for row in range(first_row, last_row + 1)]
            x_edges = [row_header]
            for width in column_widths:
                x_edges.append(x_edges[-1] + width)
            y_edges = [column_header]
            for height in row_heights:
                y_edges.append(y_edges[-1] + height)
            image = Image.new("RGB", (x_edges[-1] + 2, y_edges[-1] + 2), "white")
            draw = ImageDraw.Draw(image)
            overlay = ImageDraw.Draw(image, "RGBA")
            header_font = self._font(12 * scale, bold=True)
            draw.rectangle((0, 0, image.width, column_header), fill="#eef1f4", outline="#aeb6bf", width=1 * scale)
            draw.rectangle((0, 0, row_header, image.height), fill="#eef1f4", outline="#aeb6bf", width=1 * scale)
            for index, column in enumerate(range(first_col, last_col + 1)):
                x0, x1 = x_edges[index], x_edges[index + 1]
                label = get_column_letter(column)
                box = draw.textbbox((0, 0), label, font=header_font)
                draw.text(((x0 + x1 - (box[2] - box[0])) / 2, 5 * scale), label, fill="#4b5662", font=header_font)
                draw.line((x1, 0, x1, image.height), fill="#b9c1ca", width=1 * scale)
            for index, row in enumerate(range(first_row, last_row + 1)):
                y0, y1 = y_edges[index], y_edges[index + 1]
                label = str(row)
                box = draw.textbbox((0, 0), label, font=header_font)
                draw.text(((row_header - (box[2] - box[0])) / 2, (y0 + y1 - (box[3] - box[1])) / 2), label, fill="#4b5662", font=header_font)
                draw.line((0, y1, image.width, y1), fill="#b9c1ca", width=1 * scale)

            merged_anchor: dict[tuple[int, int], tuple[int, int, int, int]] = {}
            merged_covered: set[tuple[int, int]] = set()
            for merged_range in sheet.merged_cells.ranges:
                c0, r0, c1, r1 = merged_range.bounds
                if c1 < first_col or c0 > last_col or r1 < first_row or r0 > last_row:
                    continue
                merged_anchor[(r0, c0)] = (c0, r0, c1, r1)
                for row in range(r0, r1 + 1):
                    for column in range(c0, c1 + 1):
                        if (row, column) != (r0, c0):
                            merged_covered.add((row, column))

            def color_of(cell, default: str) -> str:
                color = getattr(cell.fill, "fgColor", None)
                rgb = getattr(color, "rgb", None)
                if isinstance(rgb, str) and len(rgb) >= 6 and cell.fill.fill_type:
                    return "#" + rgb[-6:]
                return default

            def formatted_value(cell) -> str:
                value = cell.value
                if value is None:
                    return ""
                if is_date_format(str(cell.number_format)) and hasattr(value, "strftime"):
                    return value.strftime("%Y-%m-%d")
                if isinstance(value, (int, float)):
                    pattern = str(cell.number_format or "")
                    if "%" in pattern:
                        return f"{value * 100:,.1f}%".replace(".0%", "%")
                    if "," in pattern:
                        decimals = len(pattern.rsplit(".", 1)[1]) if "." in pattern else 0
                        return f"{value:,.{decimals}f}"
                return " ".join(str(value).split())

            for row_index, row_number in enumerate(range(first_row, last_row + 1)):
                for col_index, col_number in enumerate(range(first_col, last_col + 1)):
                    if (row_number, col_number) in merged_covered:
                        continue
                    cell = sheet.cell(row_number, col_number)
                    if isinstance(cell, MergedCell):
                        continue
                    bounds = merged_anchor.get((row_number, col_number))
                    end_col = min(bounds[2], last_col) if bounds else col_number
                    end_row = min(bounds[3], last_row) if bounds else row_number
                    x0 = x_edges[col_index]
                    y0 = y_edges[row_index]
                    x1 = x_edges[end_col - first_col + 1]
                    y1 = y_edges[end_row - first_row + 1]
                    draw.rectangle((x0, y0, x1, y1), fill=color_of(cell, "#ffffff"), outline="#b5bdc7", width=1 * scale)
                    highlighted = any(
                        not (end_row < b[1] or row_number > b[3] or end_col < b[0] or col_number > b[2])
                        for b in bounds
                    )
                    if highlighted:
                        overlay.rectangle((x0, y0, x1, y1), fill=(255, 219, 72, 72), outline=(255, 153, 0, 255), width=3 * scale)
                    text = formatted_value(cell)
                    if not text:
                        continue
                    font_size = int(max(10, min(20, float(cell.font.sz or 11))) * 1.25 * scale)
                    font = self._font(font_size, bold=bool(cell.font.bold))
                    available = max(20, x1 - x0 - 16 * scale)
                    lines = []
                    current = ""
                    for character in text:
                        candidate = current + character
                        if current and draw.textlength(candidate, font=font) > available:
                            lines.append(current)
                            current = character
                        else:
                            current = candidate
                    if current:
                        lines.append(current)
                    line_height = int(font_size * 1.25)
                    max_lines = max(1, int((y1 - y0 - 10 * scale) / line_height))
                    if len(lines) > max_lines:
                        lines = lines[:max_lines]
                        lines[-1] = lines[-1][:-1] + "…" if lines[-1] else "…"
                    text_height = len(lines) * line_height
                    vertical = str(cell.alignment.vertical or "center")
                    ty = y0 + 6 * scale if vertical == "top" else y1 - text_height - 6 * scale if vertical == "bottom" else y0 + (y1 - y0 - text_height) / 2
                    horizontal = str(cell.alignment.horizontal or "general")
                    for line in lines:
                        line_width = draw.textlength(line, font=font)
                        tx = x1 - line_width - 8 * scale if horizontal == "right" else x0 + (x1 - x0 - line_width) / 2 if horizontal == "center" else x0 + 8 * scale
                        draw.text((tx, ty), line, fill="#1d2530", font=font)
                        ty += line_height
            return self._png(image)
        finally:
            workbook.close()

    @staticmethod
    def _xlsx_location(metadata: dict[str, Any]) -> tuple[str, str]:
        sheet_name = str(metadata.get("sheet_name") or "")
        cell_range = str(metadata.get("cell_range") or "")
        if sheet_name and cell_range:
            return sheet_name, cell_range
        source_location = str(metadata.get("source_location") or "")
        matched = re.fullmatch(r"sheet:(.+):row:(\d+)", source_location)
        if matched:
            return matched.group(1), f"ROW:{matched.group(2)}"
        return sheet_name, cell_range or "A1"

    def _text(self, evidence: dict[str, Any]) -> bytes:
        from PIL import Image, ImageDraw

        content = " ".join(str(evidence.get("content") or "").split())
        lines = textwrap.wrap(content, width=58)[:12] or ["원문을 표시할 수 없습니다."]
        image = Image.new("RGB", (1100, 90 + len(lines) * 34), "white")
        draw = ImageDraw.Draw(image, "RGBA")
        font = self._font(19)
        draw.rectangle((20, 20, image.width - 20, image.height - 20), fill=(255, 229, 120, 100), outline=(255, 160, 0, 255), width=5)
        for index, line in enumerate(lines):
            draw.text((38, 40 + index * 34), line, fill="#20242b", font=font)
        return self._png(image)
