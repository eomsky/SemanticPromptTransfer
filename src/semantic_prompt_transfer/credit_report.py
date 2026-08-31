from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from .config import DocumentScope
from .domain import CreditFact, ReviewItem
from .identity import evidence_id


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


@dataclass(frozen=True)
class CreditFieldMapping:
    field_id: str
    field_name: str
    sheet_name: str
    cell_range: str
    review_items: tuple[ReviewItem, ...] = ()
    common: bool = False
    unit: str | None = None
    unit_cell: str | None = None
    period: str | None = None
    period_cell: str | None = None

    @classmethod
    def from_dict(cls, value: dict[str, Any]) -> "CreditFieldMapping":
        review_items = tuple(ReviewItem(str(item)) for item in value.get("review_items", []))
        common = bool(value.get("common", False))
        if not common and not review_items:
            raise ValueError(f"field {value.get('field_id')} needs review_items or common=true")
        return cls(
            field_id=str(value["field_id"]),
            field_name=str(value.get("field_name") or value["field_id"]),
            sheet_name=str(value["sheet_name"]),
            cell_range=str(value.get("cell_range") or value.get("cell")),
            review_items=review_items,
            common=common,
            unit=value.get("unit"),
            unit_cell=value.get("unit_cell"),
            period=value.get("period"),
            period_cell=value.get("period_cell"),
        )


@dataclass(frozen=True)
class CreditReportTemplate:
    template_id: str
    version: str
    mappings: tuple[CreditFieldMapping, ...]

    @classmethod
    def from_json(cls, path: str | Path) -> "CreditReportTemplate":
        value = json.loads(Path(path).read_text(encoding="utf-8"))
        return cls(
            template_id=str(value["template_id"]),
            version=str(value["version"]),
            mappings=tuple(CreditFieldMapping.from_dict(row) for row in value["mappings"]),
        )


@dataclass(frozen=True)
class CreditReportParseResult:
    template_id: str
    template_version: str
    source_hash: str
    facts: tuple[CreditFact, ...]

    def to_dict(self) -> dict[str, Any]:
        return {
            "template_id": self.template_id,
            "template_version": self.template_version,
            "source_hash": self.source_hash,
            "facts": [fact.to_dict() for fact in self.facts],
        }


class CreditReportParser:
    """Extract mapped values from a versioned, standardized Excel workbook."""

    def parse(
        self,
        workbook_path: str | Path,
        template: CreditReportTemplate,
        scope: DocumentScope,
    ) -> CreditReportParseResult:
        if scope.document_kind != "credit_report":
            raise ValueError("credit report parsing requires document_kind='credit_report'")
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - dependency error is environment-specific
            raise RuntimeError("openpyxl is required to parse credit reports") from exc

        source = Path(workbook_path)
        values_book = load_workbook(source, data_only=True, read_only=False)
        formula_book = load_workbook(source, data_only=False, read_only=False)
        source_hash = _sha256(source)
        facts: list[CreditFact] = []

        for mapping in template.mappings:
            if mapping.sheet_name not in values_book.sheetnames:
                raise ValueError(f"missing sheet: {mapping.sheet_name}")
            values_sheet = values_book[mapping.sheet_name]
            formula_sheet = formula_book[mapping.sheet_name]
            unit = mapping.unit or (
                str(values_sheet[mapping.unit_cell].value) if mapping.unit_cell else None
            )
            period = mapping.period or (
                str(values_sheet[mapping.period_cell].value) if mapping.period_cell else None
            )
            cells = values_sheet[mapping.cell_range]
            if not isinstance(cells, tuple):
                cells = ((cells,),)
            elif cells and not isinstance(cells[0], tuple):
                cells = (cells,)
            for row in cells:
                for cell in row:
                    if cell.value is None or str(cell.value).strip() == "":
                        continue
                    coordinate = cell.coordinate
                    formula_value = formula_sheet[coordinate].value
                    formula = formula_value if isinstance(formula_value, str) and formula_value.startswith("=") else None
                    field_id = mapping.field_id if mapping.cell_range == coordinate else f"{mapping.field_id}:{coordinate}"
                    facts.append(
                        CreditFact(
                            fact_id=evidence_id(
                                "FACT",
                                scope.tenant_id,
                                scope.case_id,
                                scope.document_id,
                                field_id,
                                coordinate,
                            ),
                            field_id=field_id,
                            field_name=mapping.field_name,
                            value=cell.value,
                            unit=unit,
                            period=period,
                            review_items=mapping.review_items,
                            common=mapping.common,
                            document_id=scope.document_id,
                            source_filename=scope.source_filename or source.name,
                            sheet_name=mapping.sheet_name,
                            cell_range=coordinate,
                            formula=formula,
                            source_hash=source_hash,
                        )
                    )
        return CreditReportParseResult(
            template_id=template.template_id,
            template_version=template.version,
            source_hash=source_hash,
            facts=tuple(facts),
        )

    def parse_semantic_workbook(
        self,
        workbook_path: str | Path,
        scope: DocumentScope,
    ) -> CreditReportParseResult:
        """Extract populated rows from the seven-sheet operational credit report."""
        if scope.document_kind != "credit_report":
            raise ValueError("credit report parsing requires document_kind='credit_report'")
        from openpyxl import load_workbook

        source = Path(workbook_path)
        book = load_workbook(source, data_only=True, read_only=False)
        source_hash = _sha256(source)
        routing = {
            "1. 업체현황": ReviewItem.ordered(),
            "2. 영업현황": (ReviewItem.PROFITABILITY, ReviewItem.MAJOR_CUSTOMERS),
            "3. 재무현황": (
                ReviewItem.MAJOR_ACCOUNTS,
                ReviewItem.PROFITABILITY,
                ReviewItem.FINANCIAL_STABILITY,
                ReviewItem.CASH_FLOW,
            ),
            "4. 분식체크": (ReviewItem.FINANCIAL_STABILITY, ReviewItem.CASH_FLOW),
            "5. 추정재무현황": (
                ReviewItem.MAJOR_ACCOUNTS,
                ReviewItem.PROFITABILITY,
                ReviewItem.FINANCIAL_STABILITY,
                ReviewItem.CASH_FLOW,
            ),
            "6. 신용평가표": ReviewItem.ordered(),
            "7. 종합의견": ReviewItem.ordered(),
        }
        priority = (
            "7. 종합의견",
            "6. 신용평가표",
            "3. 재무현황",
            "5. 추정재무현황",
            "2. 영업현황",
            "4. 분식체크",
            "1. 업체현황",
        )
        missing = [name for name in routing if name not in book.sheetnames]
        if missing:
            book.close()
            raise ValueError("missing operational credit-report sheets: " + ", ".join(missing))
        facts: list[CreditFact] = []
        try:
            for sheet_name in priority:
                sheet = book[sheet_name]
                row_items = tuple(routing[sheet_name])
                for row_index, row in enumerate(sheet.iter_rows(), start=1):
                    values = [
                        f"{cell.coordinate}={str(cell.value).strip()}"
                        for cell in row
                        if cell.value is not None and str(cell.value).strip()
                    ]
                    if not values:
                        continue
                    row_text = " | ".join(values)
                    if "재무제표 주요계정(현황 및 향후전망)" in row_text:
                        row_items = (ReviewItem.MAJOR_ACCOUNTS,)
                    elif "수익성(현황 및 향후전망)" in row_text:
                        row_items = (ReviewItem.PROFITABILITY,)
                    elif "재무안정성 및 자산의 질(현황 및 향후전망)" in row_text:
                        row_items = (ReviewItem.FINANCIAL_STABILITY,)
                    elif "현금흐름 및 채무상환능력(현황 및 향후전망)" in row_text:
                        row_items = (ReviewItem.CASH_FLOW,)
                    elif "주요 매출처 및 매출비중 변동 추이" in row_text:
                        row_items = (ReviewItem.MAJOR_CUSTOMERS,)
                    elif any(
                        title in row_text
                        for title in (
                            "금융기관별 차입금 추이",
                            "과목별 차입금 추이",
                            "계정과목별 차입금 추이",
                        )
                    ):
                        row_items = (
                            ReviewItem.FINANCIAL_STABILITY,
                            ReviewItem.CASH_FLOW,
                        )
                    elif "현금흐름표" in row_text:
                        row_items = (ReviewItem.CASH_FLOW,)
                    field_id = f"{sheet_name}:row:{row_index}"
                    facts.append(
                        CreditFact(
                            fact_id=evidence_id(
                                "FACT", scope.tenant_id, scope.case_id, scope.document_id, field_id
                            ),
                            field_id=field_id,
                            field_name=f"{sheet_name} {row_index}행",
                            value=" | ".join(values),
                            unit=None,
                            period=None,
                            review_items=row_items,
                            common=False,
                            document_id=scope.document_id,
                            source_filename=scope.source_filename or source.name,
                            sheet_name=sheet_name,
                            cell_range=f"{row[0].coordinate}:{row[-1].coordinate}",
                            source_hash=source_hash,
                        )
                    )
        finally:
            book.close()
        return CreditReportParseResult(
            template_id="operational-credit-report-seven-sheet",
            template_version="1.0",
            source_hash=source_hash,
            facts=tuple(facts),
        )
