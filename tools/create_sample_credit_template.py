from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def main() -> None:
    target = (
        Path(__file__).resolve().parents[1]
        / "src/semantic_prompt_transfer/examples/operational/credit_report_sample_template.xlsx"
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "기초자료"
    sheet.append(["항목", "입력값", "기준기간"])
    sheet.append(["매출액", None, None])
    sheet.append(["영업이익", None, None])
    common = workbook.create_sheet("공통")
    common.append(["항목", "입력값"])
    common.append(["기업개요", None])
    for worksheet in workbook.worksheets:
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        worksheet.freeze_panes = "A2"
        worksheet.column_dimensions["A"].width = 22
        worksheet.column_dimensions["B"].width = 36
        worksheet.column_dimensions["C"].width = 18
    workbook.save(target)
    print(target)


if __name__ == "__main__":
    main()

