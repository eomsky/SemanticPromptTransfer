from __future__ import annotations

import json
import re
import tempfile
import threading
import unittest
from http.server import BaseHTTPRequestHandler, HTTPServer
from pathlib import Path

import numpy as np
from openpyxl import Workbook, load_workbook

from semantic_prompt_transfer import (
    CaseContext,
    CreditFieldMapping,
    CreditReportTemplate,
    DocumentKind,
    DocumentScope,
    EphemeralColabConfig,
    EphemeralColabRuntime,
    EphemeralReviewJobService,
    EvidenceCaptureService,
    EvidenceTemplateGenerator,
    FewShotExample,
    FewShotRegistry,
    FewShotSelector,
    FileStatus,
    OpenAICompatibleHttpGenerator,
    PocIdentityService,
    PocSessionManager,
    PocUploadProcessor,
    RemoteGenerationConfig,
    ReviewItem,
    ShardedAttachmentRetriever,
    ShardedNpzVectorStore,
)
from semantic_prompt_transfer._chunk_builder_base import ChunkRecord
from semantic_prompt_transfer.encoding import EncoderBackend
from semantic_prompt_transfer.registry import OperationalRegistry


class FakeEncoder(EncoderBackend):
    dimension = 8

    @staticmethod
    def _matrix(texts):
        rows = []
        for text in texts:
            row = np.zeros(8, dtype=np.float32)
            for index, char in enumerate(str(text)):
                row[(ord(char) + index) % 8] += 1
            row /= max(float(np.linalg.norm(row)), 1e-12)
            rows.append(row)
        return np.vstack(rows) if rows else np.empty((0, 8), dtype=np.float32)

    def encode_documents(self, texts):
        return self._matrix(texts)

    def encode_queries(self, texts):
        return self._matrix(texts)

    def metadata(self):
        return {"provider": "fake", "dimension": 8}


class ColabPocTests(unittest.TestCase):
    def test_v024_notebook_is_single_llm_streaming_and_fail_fast_ngrok(self):
        root = Path(__file__).resolve().parents[1]
        notebook_path = root / "notebooks/SemanticPromptTransfer_v0.24_COLAB_LAUNCHER.ipynb"
        notebook = json.loads(notebook_path.read_text(encoding="utf-8"))
        self.assertEqual(len(notebook["cells"]), 6)
        code = "\n".join(
            "".join(cell["source"])
            for cell in notebook["cells"]
            if cell["cell_type"] == "code"
        )
        for cell in notebook["cells"]:
            if cell["cell_type"] == "code":
                compile("".join(cell["source"]), str(notebook_path), "exec")
        self.assertEqual(code.count("FEW_SHOT_1 ="), 1)
        self.assertEqual(code.count("FEW_SHOT_2 ="), 1)
        self.assertEqual(code.count("FEW_SHOT_3 ="), 1)
        self.assertNotIn("TwoPassReviewGenerator", code)

    def test_v025_notebook_uses_moe_vllm_gpu_embedding_and_anonymous_scope(self):
        root = Path(__file__).resolve().parents[1]
        notebook_path = root / "notebooks/SemanticPromptTransfer_v0.25_COLAB_LAUNCHER.ipynb"
        notebook = json.loads(notebook_path.read_text(encoding="utf-8"))
        code = "\n".join(
            "".join(cell.get("source", []))
            for cell in notebook["cells"]
            if cell.get("cell_type") == "code"
        )
        for cell in notebook["cells"]:
            if cell.get("cell_type") == "code":
                compile("".join(cell.get("source", [])), notebook_path.name, "exec")
        self.assertIn("google/gemma-4-26B-A4B-it", code)
        self.assertIn("https://wheels.vllm.ai/nightly/cu129", code)
        self.assertIn('"--pre"', code)
        self.assertIn('"--max-num-seqs", "4"', code)
        self.assertIn('"--async-scheduling"', code)
        self.assertIn(
            '"--limit-mm-per-prompt", json.dumps({"image": 0, "audio": 0}',
            code,
        )
        self.assertNotIn('"image=0,audio=0"', code)
        self.assertIn("E5GpuEncoder", code)
        self.assertIn("anonymous_access=True", code)
        self.assertIn("NGROK_AUTHTOKEN", code)
        self.assertNotIn("SPT_GATE_PASSWORD", code)
        self.assertNotIn("auth=f", code)
        for number in (1, 2, 3):
            cell = "".join(notebook["cells"][number + 1]["source"])
            self.assertIn(f"FEW_SHOT_{number}", cell)
            self.assertGreater(len(cell), 2000)
        self.assertIn("max_new_tokens=700", code)
        self.assertIn('secret("NGROK_AUTHTOKEN")', code)
        self.assertLess(code.index("ngrok.connect"), code.index("loading vLLM"))

    def test_one_click_colab_notebook_has_single_executable_cell(self):
        root = Path(__file__).resolve().parents[1]
        notebook_path = (
            root / "notebooks/SemanticPromptTransfer_v0.22_COLAB_LAUNCHER.ipynb"
        )
        notebook = json.loads(notebook_path.read_text(encoding="utf-8"))
        code_cells = [
            cell for cell in notebook["cells"] if cell["cell_type"] == "code"
        ]
        self.assertEqual(len(code_cells), 1)
        code = "".join(code_cells[0]["source"])
        compile(code, str(notebook_path), "exec")
        self.assertIn("drive.mount", code)
        self.assertIn("COLAB_ASSETS.json", code)
        self.assertIn("semantic-prompt-transfer[poc]", code)
        self.assertIn("gzip.open", code)
        self.assertIn('auth=f"{gate_user}:{gate_password}"', code)
        self.assertIn("root=runtime_root", code)
        self.assertNotIn("runtime-assets/v0.22/uploads", code)

    def test_downloadable_credit_template_matches_mapping_contract(self):
        example_root = (
            Path(__file__).resolve().parents[1]
            / "src/semantic_prompt_transfer/examples/operational"
        )
        template = CreditReportTemplate.from_json(
            example_root / "credit_report_template.json"
        )
        workbook = load_workbook(
            example_root / "credit_report_sample_template.xlsx",
            data_only=False,
            read_only=False,
        )
        try:
            for mapping in template.mappings:
                self.assertIn(mapping.sheet_name, workbook.sheetnames)
                self.assertIsNotNone(workbook[mapping.sheet_name][mapping.cell_range])
                if mapping.period_cell:
                    self.assertIsNotNone(
                        workbook[mapping.sheet_name][mapping.period_cell]
                    )
        finally:
            workbook.close()

    def test_session_token_is_short_lived_and_scope_bound(self):
        now = [1000.0]
        manager = PocSessionManager(
            "test-code-123", ttl_seconds=60, clock=lambda: now[0]
        )
        with self.assertRaises(PermissionError):
            manager.create("wrong-code")
        grant = manager.create("test-code-123", label="부서 A")
        session = manager.require(
            grant.token, tenant_id=grant.session.tenant_id, case_id=grant.session.case_id
        )
        self.assertEqual(session.label, "부서 A")
        with self.assertRaises(PermissionError):
            manager.require(grant.token, case_id="another-case")
        now[0] += 61
        with self.assertRaises(PermissionError):
            manager.require(grant.token)

    def test_signup_login_uses_employee_number_and_isolates_user_case(self):
        with tempfile.TemporaryDirectory() as tmp:
            identities = PocIdentityService(Path(tmp) / "identity.sqlite")
            first = identities.register(
                department="기업금융부", name="테스트 사용자", employee_number="E10001"
            )
            self.assertEqual(first["user_id"], "E10001")
            with self.assertRaises(ValueError):
                identities.register(
                    department="기업금융부", name="중복", employee_number="E10001"
                )
            with self.assertRaises(PermissionError):
                identities.login("E10001", "wrong")

            grant = identities.login("E10001", "E10001")
            session = identities.require(
                grant.token,
                tenant_id=grant.session.tenant_id,
                case_id=grant.session.case_id,
            )
            self.assertEqual(session.department, "기업금융부")
            self.assertEqual(session.name, "테스트 사용자")
            with self.assertRaises(PermissionError):
                identities.require(grant.token, case_id="another-user-case")

            second_grant = identities.login("E10001", "E10001")
            self.assertEqual(second_grant.session.case_id, grant.session.case_id)
            identities.register(
                department="심사부", name="다른 사용자", employee_number="E10002"
            )
            other = identities.login("E10002", "E10002")
            self.assertNotEqual(other.session.case_id, grant.session.case_id)
            identities.close()

    def test_document_shards_search_and_delete_without_rewriting_other_document(self):
        with tempfile.TemporaryDirectory() as tmp:
            store = ShardedNpzVectorStore(Path(tmp) / "vectors")

            def record(document_id, chunk_id, text):
                return ChunkRecord(
                    chunk_id,
                    text,
                    text,
                    {
                        "tenant_id": "poc",
                        "case_id": "case-a",
                        "document_id": document_id,
                        "global_chunk_id": f"G-{document_id}-{chunk_id}",
                    },
                )

            encoder = FakeEncoder()
            rows_a = [record("doc-a", "1", "매출액 증가")]
            rows_b = [record("doc-b", "1", "현금흐름 개선")]
            store.upsert_document(rows_a, encoder.encode_documents(["매출액 증가"]))
            store.upsert_document(rows_b, encoder.encode_documents(["현금흐름 개선"]))
            self.assertEqual(store.count({"tenant_id": "poc", "case_id": "case-a"}), 2)
            hits = store.search(
                encoder.encode_queries(["현금흐름 개선"])[0],
                filters={"tenant_id": "poc", "case_id": "case-a"},
            )
            self.assertEqual(hits[0]["metadata"]["document_id"], "doc-b")
            removed = store.delete_document(DocumentScope("poc", "case-a", "doc-a"))
            self.assertEqual(removed, 1)
            self.assertEqual(store.count(), 1)
            self.assertEqual(len(list((Path(tmp) / "vectors").glob("*.npz"))), 1)

    def test_runtime_rejects_drive_path_and_erases_root_on_close(self):
        with self.assertRaises(ValueError):
            EphemeralColabConfig(
                root="/content/drive/MyDrive/poc", require_content_root=True
            )
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "runtime"
            runtime = EphemeralColabRuntime(
                EphemeralColabConfig(
                    root=root, require_content_root=False, clean_start=True
                )
            )
            self.assertEqual(runtime.health()["storage_mode"], "COLAB_EPHEMERAL_ONLY")
            result = runtime.close(purge=True)
            self.assertTrue(result["purged"])
            self.assertFalse(root.exists())

    def test_registry_connection_is_safe_for_background_thread_status_updates(self):
        registry = OperationalRegistry()
        registry.register_document(
            tenant_id="poc",
            case_id="case",
            document_id="doc",
            filename="a.txt",
            document_kind=DocumentKind.ATTACHMENT,
        )
        failures = []

        def update():
            try:
                registry.transition_document("poc", "case", "doc", FileStatus.VALIDATING)
            except Exception as exc:  # pragma: no cover - assertion captures it
                failures.append(exc)

        thread = threading.Thread(target=update)
        thread.start()
        thread.join()
        self.assertEqual(failures, [])
        self.assertEqual(
            registry.get_document("poc", "case", "doc").status,
            FileStatus.VALIDATING,
        )

    def test_upload_review_and_session_purge_end_to_end(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp) / "runtime"
            runtime = EphemeralColabRuntime(
                EphemeralColabConfig(
                    root=root, require_content_root=False, clean_start=True
                )
            )
            template = CreditReportTemplate(
                "test",
                "1",
                (
                    CreditFieldMapping(
                        "summary",
                        "공통 기초자료",
                        "기초자료",
                        "B2",
                        review_items=ReviewItem.ordered(),
                    ),
                ),
            )
            processor = PocUploadProcessor(
                FakeEncoder(), runtime.vectors, runtime.artifacts, credit_template=template
            )

            credit_scope = DocumentScope("poc", "case", "credit")
            credit_source = runtime.artifacts.put(credit_scope, "credit.xlsx", b"placeholder")
            workbook = Workbook()
            workbook.active.title = "기초자료"
            workbook.active["B2"] = "매출 및 현금흐름 기초자료"
            workbook.save(credit_source)
            runtime.application.register_upload(
                credit_scope,
                filename="credit.xlsx",
                document_kind=DocumentKind.CREDIT_REPORT,
                size_bytes=credit_source.stat().st_size,
                storage_uri=str(credit_source),
                derived_uri=str(runtime.artifacts.derived_path(credit_scope)),
            )
            processor.process(
                credit_scope,
                credit_source,
                DocumentKind.CREDIT_REPORT,
                lambda status, progress, message: runtime.application.update_upload(
                    credit_scope, status, progress=progress, message=message
                ),
            )

            attachment_scope = DocumentScope("poc", "case", "attachment")
            attachment_source = runtime.artifacts.put(
                attachment_scope,
                "attachment.txt",
                "주요 매출처 비중과 영업현금흐름을 확인하였다.".encode("utf-8"),
            )
            runtime.application.register_upload(
                attachment_scope,
                filename="attachment.txt",
                document_kind=DocumentKind.ATTACHMENT,
                size_bytes=attachment_source.stat().st_size,
                storage_uri=str(attachment_source),
                derived_uri=str(runtime.artifacts.derived_path(attachment_scope)),
            )
            processor.process(
                attachment_scope,
                attachment_source,
                DocumentKind.ATTACHMENT,
                lambda status, progress, message: runtime.application.update_upload(
                    attachment_scope, status, progress=progress, message=message
                ),
            )
            self.assertGreater(runtime.vectors.count(), 0)

            examples = [
                FewShotExample(
                    f"FS-{item.value}",
                    item,
                    "입력",
                    "근거에 따라 현황과 전망을 기술한다.",
                    loan_types=("*",),
                    industry_codes=("*",),
                )
                for item in ReviewItem.ordered()
            ]
            service = EphemeralReviewJobService(
                runtime,
                ShardedAttachmentRetriever(FakeEncoder(), runtime.vectors),
                FewShotSelector(FewShotRegistry(examples)),
                EvidenceTemplateGenerator(),
            )
            job = service.start("poc", "case")
            result = service.run(str(job["job_id"]))
            self.assertEqual(len(result.sections), 5)
            self.assertTrue(Path(result.output_path).is_file())
            self.assertEqual(runtime.registry.get_job(str(job["job_id"])).progress, 100)

            removed = runtime.purge_case("poc", "case")
            self.assertGreaterEqual(removed["removed_vectors"], 1)
            self.assertEqual(runtime.vectors.count(), 0)
            self.assertEqual(runtime.registry.stats()["documents"], 0)
            runtime.close(purge=True)

    def test_review_job_streams_five_sections_and_captures_credit_evidence(self):
        class StreamingCitationGenerator:
            def stream(self, messages):
                match = re.search(r"evidence_id=([^\n]+)", messages[1]["content"])
                yield "확인된 기초자료를 근거로 "
                yield f"현황과 향후 전망을 검토하였다. [{match.group(1)}]"

            def generate(self, messages):
                return "".join(self.stream(messages))

        with tempfile.TemporaryDirectory() as tmp:
            runtime = EphemeralColabRuntime(
                EphemeralColabConfig(
                    root=Path(tmp) / "runtime",
                    require_content_root=False,
                    clean_start=True,
                )
            )
            template = CreditReportTemplate(
                "stream-test",
                "1",
                (
                    CreditFieldMapping(
                        "summary",
                        "심사 기초자료",
                        "기초자료",
                        "B2",
                        review_items=ReviewItem.ordered(),
                    ),
                ),
            )
            processor = PocUploadProcessor(
                FakeEncoder(), runtime.vectors, runtime.artifacts, credit_template=template
            )
            scope = DocumentScope("poc", "case", "credit")
            source = runtime.artifacts.put(scope, "credit.xlsx", b"placeholder")
            workbook = Workbook()
            workbook.active.title = "기초자료"
            workbook.active["B2"] = "신용조사서 우선 근거"
            workbook.save(source)
            runtime.application.register_upload(
                scope,
                filename="credit.xlsx",
                document_kind=DocumentKind.CREDIT_REPORT,
                size_bytes=source.stat().st_size,
                storage_uri=str(source),
                derived_uri=str(runtime.artifacts.derived_path(scope)),
            )
            processor.process(
                scope,
                source,
                DocumentKind.CREDIT_REPORT,
                lambda status, progress, message: runtime.application.update_upload(
                    scope, status, progress=progress, message=message
                ),
            )
            service = EphemeralReviewJobService(
                runtime,
                ShardedAttachmentRetriever(FakeEncoder(), runtime.vectors),
                FewShotSelector(FewShotRegistry([])),
                StreamingCitationGenerator(),
            )
            job = service.start("poc", "case")
            result = service.run(str(job["job_id"]))
            events = list(service.stream_events(str(job["job_id"])))
            self.assertEqual([event["type"] for event in events].count("section_complete"), 5)
            self.assertEqual([event["type"] for event in events].count("token"), 10)
            self.assertEqual(events[-1]["type"], "complete")
            self.assertEqual([section.review_item for section in result.sections], list(ReviewItem.ordered()))
            evidence_id = result.sections[0].evidence_ids[0]
            metadata = service.get_evidence(str(job["job_id"]), evidence_id)
            capture = service.capture_evidence(str(job["job_id"]), evidence_id)
            self.assertEqual(metadata["kind"], "xlsx")
            self.assertEqual(metadata["source_tier"], 1)
            self.assertTrue(capture.startswith(b"\x89PNG\r\n\x1a\n"))
            runtime.close(purge=True)

    def test_review_job_uses_attachment_only_rag_without_credit_report(self):
        with tempfile.TemporaryDirectory() as tmp:
            runtime = EphemeralColabRuntime(
                EphemeralColabConfig(
                    root=Path(tmp) / "runtime",
                    require_content_root=False,
                    clean_start=True,
                )
            )
            encoder = FakeEncoder()
            processor = PocUploadProcessor(
                encoder,
                runtime.vectors,
                runtime.artifacts,
            )
            scope = DocumentScope("anonymous-client", "case", "business-report")
            source = runtime.artifacts.put(
                scope,
                "business-report.txt",
                (
                    "매출액과 영업이익이 증가하였다. 재고자산과 매입채무가 변동하였다. "
                    "영업현금흐름과 차입금 상환능력을 점검하였다. 주요 매출처 비중이 변동하였다."
                ).encode("utf-8"),
            )
            runtime.application.register_upload(
                scope,
                filename="business-report.txt",
                document_kind=DocumentKind.ATTACHMENT,
                size_bytes=source.stat().st_size,
                storage_uri=str(source),
                derived_uri=str(runtime.artifacts.derived_path(scope)),
            )
            processor.process(
                scope,
                source,
                DocumentKind.ATTACHMENT,
                lambda status, progress, message: runtime.application.update_upload(
                    scope, status, progress=progress, message=message
                ),
            )
            service = EphemeralReviewJobService(
                runtime,
                ShardedAttachmentRetriever(encoder, runtime.vectors),
                FewShotSelector(FewShotRegistry([])),
                EvidenceTemplateGenerator(),
            )
            job = service.start("anonymous-client", "case")
            result = service.run(str(job["job_id"]))
            self.assertEqual(len(result.sections), 5)
            for section, prompt in zip(result.sections, result.prompts):
                self.assertTrue(section.evidence_ids)
                self.assertTrue(all(value.startswith("ATT_") for value in section.evidence_ids))
                self.assertFalse(prompt.manifest["credit_report_available"])
                self.assertTrue(prompt.manifest["attachment_evidence_available"])
            runtime.close(purge=True)

    def test_pdf_capture_uses_page_block_coordinates(self):
        try:
            import fitz
        except ImportError:  # pragma: no cover
            self.skipTest("PyMuPDF is not installed")
        with tempfile.TemporaryDirectory() as tmp:
            runtime = EphemeralColabRuntime(
                EphemeralColabConfig(
                    root=Path(tmp) / "runtime",
                    require_content_root=False,
                    clean_start=True,
                )
            )
            scope = DocumentScope("poc", "case", "pdf")
            source = runtime.root / "uploads" / "poc" / "case" / "pdf" / "evidence.pdf"
            source.parent.mkdir(parents=True, exist_ok=True)
            document = fitz.open()
            page = document.new_page(width=400, height=300)
            page.insert_text((45, 90), "Evidence source line")
            document.save(source)
            document.close()
            runtime.registry.register_document(
                tenant_id="poc",
                case_id="case",
                document_id="pdf",
                filename="evidence.pdf",
                document_kind=DocumentKind.ATTACHMENT,
                storage_uri=str(source),
            )
            capture = EvidenceCaptureService(runtime).capture_png(
                "poc",
                "case",
                {
                    "document_id": "pdf",
                    "page": 1,
                    "content": "Evidence source line",
                    "metadata": {"bbox": [40, 70, 220, 105]},
                },
            )
            self.assertTrue(capture.startswith(b"\x89PNG\r\n\x1a\n"))
            runtime.close(purge=True)

    def test_remote_llm_adapter_calls_and_streams_openai_compatible_endpoint(self):
        received = []

        class Handler(BaseHTTPRequestHandler):
            def do_POST(self):
                length = int(self.headers["Content-Length"])
                payload = json.loads(self.rfile.read(length))
                received.append(payload)
                if payload.get("stream"):
                    body = (
                        'data: {"choices":[{"delta":{"content":"근거 기반 "}}]}\n\n'
                        'data: {"choices":[{"delta":{"content":"스트림"}}]}\n\n'
                        "data: [DONE]\n\n"
                    ).encode("utf-8")
                    self.send_response(200)
                    self.send_header("Content-Type", "text/event-stream")
                    self.send_header("Content-Length", str(len(body)))
                    self.end_headers()
                    self.wfile.write(body)
                    return
                body = json.dumps(
                    {"choices": [{"message": {"content": "근거 기반 응답 EV-1"}}]}
                ).encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)

            def log_message(self, format, *args):
                return

        server = HTTPServer(("127.0.0.1", 0), Handler)
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        try:
            generator = OpenAICompatibleHttpGenerator(
                RemoteGenerationConfig(
                    base_url=f"http://127.0.0.1:{server.server_port}",
                    model="test-model",
                    allow_insecure_http=True,
                )
            )
            text = generator.generate([{"role": "user", "content": "작성"}])
            self.assertEqual(text, "근거 기반 응답 EV-1")
            self.assertEqual("".join(generator.stream([{"role": "user", "content": "작성"}])), "근거 기반 스트림")
            self.assertEqual(received[0]["model"], "test-model")
            self.assertFalse(received[0]["stream"])
            self.assertTrue(received[1]["stream"])
        finally:
            server.shutdown()
            thread.join()
            server.server_close()

    def test_html_declares_anonymous_runtime_and_evidence_contract(self):
        html = (
            Path(__file__).resolve().parents[1]
            / "src/semantic_prompt_transfer/examples/operational/credit_review_upload_demo.html"
        ).read_text(encoding="utf-8")
        for value in (
            "/api/v1/runtime/health",
            "/api/v1/templates/credit-report.xlsx",
            "sptAnonymousClientV1",
            "로그인 없는 시간 제한형 Colab POC",
            "Colab 연결됨",
            "심사의견.docx",
            "양식 다운로드",
            "file-x",
            "/stream?after=",
            "/capture.png",
            "심사의견 생성 내용",
            "currentStage",
        ):
            self.assertIn(value, html)
        self.assertNotIn("업로드 자료현황", html)
        self.assertNotIn("추가 산출물", html)
        self.assertNotIn("generate_optional_artifacts", html)
        self.assertNotIn("/api/v1/poc/login", html)
        self.assertNotIn("X-POC-Token", html)


if __name__ == "__main__":
    unittest.main()
